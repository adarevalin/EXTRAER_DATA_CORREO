# =========================================================
# 📦 IMPORTACIÓN DE LIBRERÍAS
# =========================================================
import os, sys                   # Manejo de rutas y entorno del sistema
import win32com.client           # Comunicación con Outlook (vía COM)
import extract_msg               # Para abrir archivos .msg (correos de Outlook)
import pdfplumber                # Para leer texto de archivos PDF
import re                        # Expresiones regulares (búsqueda de patrones)
from openpyxl import Workbook, load_workbook  # Lectura y escritura de Excel
from datetime import datetime    # Manejo de fechas
import time                      # Pausas breves para evitar errores de acceso

# =========================================================
# 📂 CONFIGURACIÓN DE RUTAS Y ARCHIVOS
# =========================================================

# Detecta si el script está empaquetado como .exe o se ejecuta en Python normal
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Carpeta donde se guardarán los PDF extraídos
PDF_DIR = os.path.join(BASE_DIR, "pdfs")
os.makedirs(PDF_DIR, exist_ok=True)

# Ruta del archivo Excel donde se guardarán las órdenes
excel_file = os.path.join(BASE_DIR, "ordenes_compra.xlsx")

# =========================================================
# 📧 CONEXIÓN A OUTLOOK
# =========================================================
try:
    # Conexión al cliente de Outlook mediante COM
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
except Exception as e:
    print("❌ Error al conectarse a Outlook:", e)
    sys.exit()  # Finaliza el programa si no se logra conectar

# =========================================================
# 📊 CREACIÓN O CARGA DEL ARCHIVO EXCEL
# =========================================================
if not os.path.exists(excel_file):
    # Si el archivo no existe, se crea uno nuevo con los encabezados
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordenes"
    ws.append([
        "Buzón", "Número de Orden", "Fecha Creación",
        "Fecha Entrega", "Sección", "Total sin IVA",
        "Total con IVA", "Descripción final"
    ])
    wb.save(excel_file)

# Carga el archivo Excel existente
wb = load_workbook(excel_file)
ws = wb.active

# Crea un conjunto con los números de orden ya guardados (para evitar duplicados)
ordenes_existentes = {str(c.value) for c in ws["B"] if c.value}

# =========================================================
# 🔍 FILTRO DE CORREOS A PROCESAR
# =========================================================
# Busca solo correos que tengan “Cencosud: Orden de Compra” en el asunto
filtro = "@SQL=\"urn:schemas:mailheader:subject\" like '%Cencosud: Orden de Compra%'"

# =========================================================
# 🔄 RECORRER TODOS LOS BUZONES DISPONIBLES EN OUTLOOK
# =========================================================
for store in outlook.Folders:
    print(f"📂 Revisando buzón: {store.Name}")

    try:
        # Accede a la bandeja de entrada del buzón actual
        inbox = store.Folders["Bandeja de entrada"]

        # Aplica el filtro para traer solo correos relevantes
        correos = inbox.Items.Restrict(filtro)
        correos.Sort("[ReceivedTime]", True)  # Ordena por fecha descendente

        # Limita el número de correos a procesar (para no sobrecargar)
        correos_a_procesar = list(correos)[:700]
    except Exception as e:
        print(f"⚠️ No se pudo acceder a la bandeja de entrada de {store.Name}: {e}")
        continue  # Si falla un buzón, pasa al siguiente

    # =====================================================
    # 📩 PROCESAR CADA CORREO ENCONTRADO
    # =====================================================
    for mail in correos_a_procesar:
        try:
            for attachment in mail.Attachments:
                # Solo procesa adjuntos con extensión .msg
                if not attachment.FileName.lower().endswith(".msg"):
                    continue

                # Guarda el archivo .msg temporalmente
                msg_path = os.path.join(BASE_DIR, attachment.FileName)
                attachment.SaveAsFile(msg_path)

                # Abre el archivo .msg con extract_msg
                msg = extract_msg.Message(msg_path)

                # =====================================================
                # 📎 PROCESAR LOS ARCHIVOS PDF DENTRO DEL .MSG
                # =====================================================
                for att in msg.attachments:
                    if not att.longFilename.lower().endswith(".pdf"):
                        continue

                    # Guarda el PDF dentro de la carpeta /pdfs
                    pdf_path = os.path.join(PDF_DIR, f"{store.Name}_{att.longFilename}")
                    with open(pdf_path, "wb") as f:
                        f.write(att.data)

                    # =====================================================
                    # 📄 EXTRAER TEXTO DEL PDF
                    # =====================================================
                    texto = ""
                    with pdfplumber.open(pdf_path) as pdf:
                        for page in pdf.pages:
                            if page.extract_text():
                                texto += page.extract_text() + "\n"

                    # =====================================================
                    # 🔎 EXTRAER DATOS CLAVE DEL TEXTO
                    # =====================================================

                    # Función auxiliar para formatear fechas
                    def parse_fecha(t):
                        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
                            try:
                                return datetime.strptime(t, fmt).strftime("%d/%m/%Y")
                            except:
                                continue
                        return t or "NO ENCONTRADO"

                    # Número de orden
                    numero_orden = re.search(r"Número\s*(?:de\s*)?orden\s*[:\s]+(\d+)", texto, re.I)
                    numero_orden = numero_orden.group(1) if numero_orden else "NO ENCONTRADO"

                    # Fecha de creación
                    fecha_creacion = re.search(r"Fecha\s*creaci[oó]n\s*[:\s]+([\d./-]+)", texto, re.I)
                    fecha_creacion = parse_fecha(fecha_creacion.group(1)) if fecha_creacion else "NO ENCONTRADO"

                    # Fecha de entrega
                    fecha_entrega = re.search(r"Fecha\s*entrega\s*[:\s]+([\d./-]+)", texto, re.I)
                    fecha_entrega = parse_fecha(fecha_entrega.group(1)) if fecha_entrega else "NO ENCONTRADO"

                    # Sección
                    match_seccion = re.search(r"Sección\s*:\s*(.+)", texto, re.I)
                    seccion = match_seccion.group(1).strip() if match_seccion else "NO ENCONTRADO"

                    # Total sin IVA
                    neto = re.search(r"Neto total sin IVA\s+[A-Z]{3}\s+([\d.,]+)", texto)
                    total_sin_iva = neto.group(1) if neto else None

                    # Total con IVA
                    totales = re.findall(r"Total\s+([\d.,]+)", texto)
                    total_con_iva = totales[-1] if totales else None

                    # Descripción final (línea siguiente a “Total”)
                    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
                    descripcion_final = next(
                        (lineas[i+1] for i, l in enumerate(lineas)
                         if l.startswith("Total") and i+1 < len(lineas)),
                        None
                    )

                    # =====================================================
                    # 💾 GUARDAR DATOS EN EXCEL (SI NO EXISTEN)
                    # =====================================================
                    if numero_orden not in ordenes_existentes:
                        ws.append([
                            store.Name, numero_orden, fecha_creacion,
                            fecha_entrega, seccion, total_sin_iva,
                            total_con_iva, descripcion_final
                        ])
                        ordenes_existentes.add(numero_orden)
                        print(f"   ✅ Guardada orden {numero_orden}")
                        wb.save(excel_file)
                    else:
                        print(f"   ⏭️ Orden {numero_orden} ya existe, saltada.")

                # Cierra el archivo .msg y elimina el temporal
                msg.close()
                try:
                    os.remove(msg_path)
                except PermissionError:
                    time.sleep(1)
                    os.remove(msg_path)

        except Exception as e:
            print(f"⚠️ Error procesando correo en {store.Name}: {e}")

# =========================================================
# 🏁 FINALIZACIÓN DEL PROCESO
# =========================================================
wb.save(excel_file)
print("✅ Proceso completado y Excel actualizado.")



